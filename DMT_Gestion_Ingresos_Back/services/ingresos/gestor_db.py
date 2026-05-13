from collections import defaultdict
from io import BytesIO
from datetime import date, datetime, time
from typing import Dict, Iterable, Optional

from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from services.db.models import Ingreso, IngresoImagen
from services.db.session import SessionLocal
from services.catalogos import CatalogosStore
from services.storage.imagenes import guardar_imagenes_ingreso, leer_imagen_base64, validar_jpeg_base64


MESES = {
    1: ("Enero", "ENE"),
    2: ("Febrero", "FEB"),
    3: ("Marzo", "MAR"),
    4: ("Abril", "ABR"),
    5: ("Mayo", "MAY"),
    6: ("Junio", "JUN"),
    7: ("Julio", "JUL"),
    8: ("Agosto", "AGO"),
    9: ("Septiembre", "SEP"),
    10: ("Octubre", "OCT"),
    11: ("Noviembre", "NOV"),
    12: ("Diciembre", "DIC"),
}


def _parse_fecha(fecha: Optional[str]) -> date:
    if not fecha:
        return date.today()
    if "/" in fecha:
        return datetime.strptime(fecha, "%d/%m/%Y").date()
    return datetime.strptime(fecha, "%Y-%m-%d").date()


def _parse_hora(valor: Optional[str]) -> time:
    if not valor:
        return datetime.now().time().replace(microsecond=0)
    return datetime.strptime(valor, "%H:%M:%S").time()


def _hora_str(valor: Optional[time]) -> str:
    return valor.strftime("%H:%M:%S") if valor else ""


def _fecha_str(valor: date) -> str:
    return valor.strftime("%d/%m/%Y")


class GestorIngresosDB:
    def __init__(self, tipo: str):
        if tipo not in {"peatonal", "vehicular"}:
            raise ValueError("tipo debe ser peatonal o vehicular")
        self.tipo = tipo

    def crear_ingreso(
        self,
        numero_cedula: str,
        nombres: str,
        apellidos: str,
        hora_entrada: str,
        departamento: str,
        motivo: str,
        imagen_usuario_base64: str,
        imagen_cedula_base64: str,
        imagen_placa_base64: Optional[str] = None,
    ) -> dict:
        validacion = self._validar_datos(
            numero_cedula, nombres, apellidos, hora_entrada, departamento, motivo
        )
        if validacion:
            return {"error": validacion}
        if self.tipo == "vehicular":
            cupo_error = self._validar_cupo_vehicular(departamento)
            if cupo_error:
                return {"error": cupo_error}

        imagenes = {
            "usuario": imagen_usuario_base64,
            "cedula": imagen_cedula_base64,
        }
        if self.tipo == "vehicular":
            imagenes["placa"] = imagen_placa_base64 or ""

        for tipo_imagen, contenido in imagenes.items():
            ok, mensaje = validar_jpeg_base64(contenido, f"Imagen {tipo_imagen}")
            if not ok:
                return {"error": mensaje}

        fecha = date.today()
        hora = _parse_hora(hora_entrada)

        with SessionLocal() as db:
            ticket = self._generar_ticket(db, fecha)
            fecha_path = f"{fecha.year}/{MESES[fecha.month][0]}/{fecha.day:02d}"
            rutas = guardar_imagenes_ingreso(self.tipo, fecha_path, ticket, imagenes)

            ingreso = Ingreso(
                ticket=ticket,
                tipo=self.tipo,
                numero_cedula=numero_cedula.strip(),
                nombres=nombres.strip(),
                apellidos=apellidos.strip(),
                departamento=departamento.strip(),
                motivo=motivo.strip(),
                fecha_ingreso=fecha,
                hora_entrada=hora,
                estado="activo",
            )
            db.add(ingreso)
            db.flush()

            for tipo_imagen, ruta in rutas.items():
                db.add(IngresoImagen(ingreso_id=ingreso.id, tipo=tipo_imagen, ruta_archivo=ruta))

            db.commit()
            db.refresh(ingreso)

            return {
                "exito": True,
                "ticket": ingreso.ticket,
                "fecha_registro": _fecha_str(ingreso.fecha_ingreso),
                "hora_entrada": _hora_str(ingreso.hora_entrada),
            }

    def leer_ingreso(self, ticket: str) -> dict:
        with SessionLocal() as db:
            ingreso = self._buscar_por_ticket(db, ticket)
            if not ingreso:
                return {"error": f"Ticket {ticket} no encontrado"}
            return self._serializar(ingreso, incluir_imagenes=True)

    def actualizar_ingreso(self, ticket: str, datos: dict) -> dict:
        permitidos = {"numero_cedula", "nombres", "apellidos", "departamento", "motivo"}
        datos_limpios = {k: v for k, v in datos.items() if k in permitidos and v is not None}
        if not datos_limpios:
            return {"error": "No hay campos válidos para actualizar"}

        with SessionLocal() as db:
            ingreso = self._buscar_por_ticket(db, ticket)
            if not ingreso:
                return {"error": f"Ticket {ticket} no encontrado"}
            for campo, valor in datos_limpios.items():
                setattr(ingreso, campo, str(valor).strip())
            db.commit()
            return {"exito": True, "ticket": ticket, "mensaje": "Registro actualizado exitosamente"}

    def registrar_salida(self, ticket: str, hora_salida: Optional[str] = None) -> dict:
        with SessionLocal() as db:
            ingreso = self._buscar_por_ticket(db, ticket)
            if not ingreso:
                return {"error": f"Ticket {ticket} no encontrado"}
            if ingreso.estado == "salida":
                return {"error": "El ingreso ya tiene salida registrada"}
            ingreso.hora_salida = _parse_hora(hora_salida)
            ingreso.estado = "salida"
            db.commit()
            return {
                "exito": True,
                "ticket": ticket,
                "hora_salida": _hora_str(ingreso.hora_salida),
                "estado": ingreso.estado,
            }

    def listar_ingresos_por_dia(self, fecha_str: str) -> dict:
        try:
            fecha = _parse_fecha(fecha_str)
        except ValueError:
            return {"error": "Formato fecha inválido. Use DD/MM/YYYY o YYYY-MM-DD"}

        with SessionLocal() as db:
            ingresos = db.execute(
                select(Ingreso)
                .where(Ingreso.tipo == self.tipo, Ingreso.fecha_ingreso == fecha)
                .order_by(Ingreso.created_at.desc(), Ingreso.id.desc())
            ).scalars().all()

            tickets = [self._serializar(ingreso, incluir_imagenes=False) for ingreso in ingresos]
            return {
                "exito": True,
                "fecha": fecha_str,
                "cantidad_tickets": len(tickets),
                "tickets": tickets,
            }

    @staticmethod
    def buscar_persona_por_cedula(numero_cedula: str) -> dict:
        cedula = "".join(ch for ch in str(numero_cedula or "") if ch.isdigit())[-10:]
        if len(cedula) != 10:
            return {"encontrado": False, "cedula": cedula}

        with SessionLocal() as db:
            ingreso = db.execute(
                select(Ingreso)
                .where(Ingreso.numero_cedula == cedula)
                .order_by(Ingreso.created_at.desc(), Ingreso.id.desc())
            ).scalars().first()

            if not ingreso:
                return {"encontrado": False, "cedula": cedula}

            return {
                "encontrado": True,
                "cedula": ingreso.numero_cedula,
                "nombres": ingreso.nombres,
                "apellidos": ingreso.apellidos,
                "ultimo_ticket": ingreso.ticket,
                "fecha_ingreso": _fecha_str(ingreso.fecha_ingreso),
            }

    @staticmethod
    def listar_todos_por_dia(fecha_str: str) -> dict:
        try:
            fecha = _parse_fecha(fecha_str)
        except ValueError:
            return {"error": "Formato fecha inválido. Use DD/MM/YYYY o YYYY-MM-DD"}

        with SessionLocal() as db:
            ingresos = db.execute(
                select(Ingreso)
                .where(Ingreso.fecha_ingreso == fecha)
                .order_by(Ingreso.created_at.desc(), Ingreso.id.desc())
            ).scalars().all()
            tickets = [GestorIngresosDB._serializar_static(ingreso, incluir_imagenes=False) for ingreso in ingresos]
            return {"exito": True, "fecha": fecha_str, "cantidad_tickets": len(tickets), "tickets": tickets}

    @staticmethod
    def listar_filtrado(
        fecha: Optional[str] = None,
        fecha_desde: Optional[str] = None,
        fecha_hasta: Optional[str] = None,
        tipo: Optional[str] = None,
        numero_cedula: Optional[str] = None,
        departamento: Optional[str] = None,
        motivo: Optional[str] = None,
        estado: Optional[str] = None,
    ) -> list[dict]:
        stmt = select(Ingreso).order_by(Ingreso.fecha_ingreso.desc(), Ingreso.created_at.desc(), Ingreso.id.desc())

        if fecha:
            fecha_exacta = _parse_fecha(fecha)
            stmt = stmt.where(Ingreso.fecha_ingreso == fecha_exacta)
        else:
            if fecha_desde:
                stmt = stmt.where(Ingreso.fecha_ingreso >= _parse_fecha(fecha_desde))
            if fecha_hasta:
                stmt = stmt.where(Ingreso.fecha_ingreso <= _parse_fecha(fecha_hasta))

        if tipo in {"peatonal", "vehicular"}:
            stmt = stmt.where(Ingreso.tipo == tipo)
        if numero_cedula:
            stmt = stmt.where(Ingreso.numero_cedula.ilike(f"%{numero_cedula.strip()}%"))
        if departamento:
            stmt = stmt.where(Ingreso.departamento == departamento)
        if motivo:
            stmt = stmt.where(Ingreso.motivo == motivo)
        if estado in {"activo", "salida"}:
            stmt = stmt.where(Ingreso.estado == estado)

        with SessionLocal() as db:
            ingresos = db.execute(stmt).scalars().all()
            return [GestorIngresosDB._serializar_static(ingreso, incluir_imagenes=False) for ingreso in ingresos]

    @staticmethod
    def exportar_excel(**filtros) -> BytesIO:
        registros = GestorIngresosDB.listar_filtrado(**filtros)
        wb = Workbook()
        ws = wb.active
        ws.title = "Ingresos"

        headers = [
            "Ticket",
            "Tipo",
            "Fecha",
            "Cedula",
            "Nombres",
            "Apellidos",
            "Departamento",
            "Motivo",
            "Hora entrada",
            "Hora salida",
            "Estado",
        ]
        ws.append(headers)

        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for item in registros:
            ws.append([
                item["ticket"],
                item["tipo"],
                item["fecha_ingreso"],
                item["numero_cedula"],
                item["nombres"],
                item["apellidos"],
                item["departamento"],
                item["motivo"],
                item["hora_entrada"],
                item["hora_salida"],
                item["estado"],
            ])

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 34)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def metricas(fecha_str: Optional[str] = None) -> dict:
        fecha = _parse_fecha(fecha_str)
        with SessionLocal() as db:
            ingresos = db.execute(select(Ingreso).where(Ingreso.fecha_ingreso == fecha)).scalars().all()

        por_tipo = {"peatonal": [], "vehicular": []}
        for ingreso in ingresos:
            por_tipo[ingreso.tipo].append(ingreso)

        return {
            "fecha": fecha.isoformat(),
            "total": len(ingresos),
            "peatonal": GestorIngresosDB._resumen(por_tipo["peatonal"]),
            "vehicular": GestorIngresosDB._resumen(por_tipo["vehicular"]),
        }

    def _validar_datos(self, *valores: str) -> Optional[str]:
        nombres_campos = ["Cédula", "Nombres", "Apellidos", "Hora de entrada", "Departamento", "Motivo"]
        for nombre, valor in zip(nombres_campos, valores):
            if not valor or not str(valor).strip():
                return f"{nombre} es requerido"
        return None

    def _generar_ticket(self, db: Session, fecha: date) -> str:
        _, abrev_mes = MESES[fecha.month]
        prefijo_tipo = "P" if self.tipo == "peatonal" else "V"
        base = f"TICKET-{prefijo_tipo}-{abrev_mes}-{fecha.day:02d}"
        total = db.execute(
            select(func.count(Ingreso.id)).where(
                Ingreso.tipo == self.tipo,
                Ingreso.fecha_ingreso == fecha,
            )
        ).scalar_one()
        return f"{base}-{total + 1:03d}"

    def _buscar_por_ticket(self, db: Session, ticket: str) -> Optional[Ingreso]:
        return db.execute(
            select(Ingreso)
            .options(selectinload(Ingreso.imagenes))
            .where(Ingreso.tipo == self.tipo, Ingreso.ticket == ticket)
        ).scalar_one_or_none()

    def _serializar(self, ingreso: Ingreso, incluir_imagenes: bool) -> dict:
        return self._serializar_static(ingreso, incluir_imagenes)

    @staticmethod
    def _serializar_static(ingreso: Ingreso, incluir_imagenes: bool) -> dict:
        data = {
            "exito": True,
            "ticket": ingreso.ticket,
            "tipo": ingreso.tipo,
            "fecha_ingreso": _fecha_str(ingreso.fecha_ingreso),
            "numero_cedula": ingreso.numero_cedula,
            "nombres": ingreso.nombres,
            "apellidos": ingreso.apellidos,
            "departamento": ingreso.departamento,
            "motivo": ingreso.motivo,
            "hora_entrada": _hora_str(ingreso.hora_entrada),
            "hora_salida": _hora_str(ingreso.hora_salida),
            "estado": ingreso.estado,
        }
        if incluir_imagenes:
            for imagen in ingreso.imagenes:
                data[f"imagen_{imagen.tipo}_base64"] = leer_imagen_base64(imagen.ruta_archivo)
        return data

    @staticmethod
    def _resumen(ingresos: Iterable[Ingreso]) -> dict:
        ingresos = list(ingresos)
        departamentos: Dict[str, dict] = defaultdict(lambda: {"total": 0, "activos": 0, "salidos": 0})
        for ingreso in ingresos:
            dept = departamentos[ingreso.departamento]
            dept["total"] += 1
            if ingreso.estado == "activo":
                dept["activos"] += 1
            else:
                dept["salidos"] += 1

        return {
            "total": len(ingresos),
            "activos": sum(1 for ingreso in ingresos if ingreso.estado == "activo"),
            "salidos": sum(1 for ingreso in ingresos if ingreso.estado == "salida"),
            "departamentos": [
                {"departamento": departamento, **valores}
                for departamento, valores in sorted(departamentos.items())
            ],
        }

    @staticmethod
    def cupos_vehiculares() -> dict:
        departamentos = CatalogosStore().listar_departamentos(solo_activos=True)
        with SessionLocal() as db:
            resultado = []
            for departamento in departamentos:
                cupo = departamento.get("cupo_vehicular")
                reset_at = GestorIngresosDB._parse_reset_at(departamento.get("parqueo_reset_at"))
                stmt = select(func.count(Ingreso.id)).where(
                    Ingreso.tipo == "vehicular",
                    Ingreso.departamento == departamento["nombre"],
                    Ingreso.estado == "activo",
                )
                if reset_at:
                    stmt = stmt.where(Ingreso.created_at >= reset_at)
                ocupados = db.execute(stmt).scalar_one()
                resultado.append({
                    "departamento_id": departamento["id"],
                    "departamento": departamento["nombre"],
                    "cupo_vehicular": cupo,
                    "ocupados": ocupados,
                    "disponibles": None if cupo is None else max(cupo - ocupados, 0),
                    "lleno": False if cupo is None else ocupados >= cupo,
                    "parqueo_reset_at": departamento.get("parqueo_reset_at"),
                })
        return {"cupos": resultado}

    @staticmethod
    def _parse_reset_at(valor: Optional[str]) -> Optional[datetime]:
        if not valor:
            return None
        try:
            return datetime.fromisoformat(valor)
        except ValueError:
            return None

    @staticmethod
    def _validar_cupo_vehicular(departamento_nombre: str) -> Optional[str]:
        cupos = GestorIngresosDB.cupos_vehiculares()["cupos"]
        for item in cupos:
            if item["departamento"] == departamento_nombre and item["cupo_vehicular"] is not None and item["lleno"]:
                return f"Cupo vehicular lleno para {departamento_nombre}. Ocupados: {item['ocupados']} de {item['cupo_vehicular']}"
        return None
